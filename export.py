"""Export website data from the WC 2026 main model.

Reads ONLY the main model workbook (not the individual submission files):
  - `_Setup_PowerQuery` A7-table  -> leaderboard.json   (official standings)
  - `_Setup_PowerQuery` row-100 flat table (one row per player, all picks)
  - `Backend` sheet               -> fixtures, results, per-team progression
  - `Scoring` sheet               -> bonus question texts + point values
  - `Results` sheet BO/BP cols    -> bonus answer key + status (hand-maintained)

Outputs:
  leaderboard.json  summary standings (+ rank movement vs previous publish)
  games.json        per-game predicted-scoreline distributions + points
  stats.json        per-team predicted progression + bonus question status

Scoring (validated against the model's own totals):
  group game: exact score = 3, correct 1X2 = 1, wrong = 0
  group winner = 3; KO "team reaches round": R32=3 R16=6 QF=9 SF=12 Final=15,
  champion = 30, 3rd place = 15.
All readers are defensive about missing results: at tournament start the
Results sheet is empty -> games show as upcoming, no winners, bonus mostly TBD.
"""

import datetime
import json
import os
import openpyxl

WORKBOOK = "WC 2026 Main model_vOPX.xlsx"
SHEET = "_Setup_PowerQuery"
OUTPUT = "leaderboard.json"
BASELINE_FILE = "daily_baseline.json"
GAMES_OUTPUT = "games.json"
STATS_OUTPUT = "stats.json"
PLAYERS_OUTPUT = "players.json"
HISTORY_OUTPUT = "history.json"
BRACKET_OUTPUT = "bracket.json"
ACH_OUTPUT = "achievements.json"

FLAT_HEADER_ROW = 100          # header row of the per-player flat picks table
STAGES = ["Group stage", "R32", "R16", "QF", "SF", "Final", "Champion"]


def sign(a, b):
    d = a - b
    return 0 if d == 0 else (1 if d > 0 else -1)


def game_points(phg, pag, hg, ag):
    if phg == hg and pag == ag:
        return 3
    if sign(phg, pag) == sign(hg, ag):
        return 1
    return 0


def as_int(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str) and v.strip().lstrip("-").isdigit():
        return int(v.strip())
    return None


# Read from a temp copy so this works even while the model is open in Excel
# (a direct open would hit a PermissionError on the locked file). The copy
# reflects the last SAVED state, which is exactly what we want to export.
import shutil
import tempfile
_tmp_model = shutil.copy(WORKBOOK, tempfile.mktemp(suffix=".xlsx"))
wb = openpyxl.load_workbook(_tmp_model, data_only=True)
ws = wb[SHEET]
bk = wb["Backend"]

# ───────────────────────── leaderboard.json ─────────────────────────

# Previous export (most recent prior state) — used to seed a new day's baseline.
prev_state = {}                # name -> {"rank":, "total":}
if os.path.exists(OUTPUT):
    try:
        with open(OUTPUT, encoding="utf-8") as f:
            for r in json.load(f).get("rows", []):
                if r.get("player") is not None:
                    prev_state[r["player"]] = {"rank": r.get("rank"), "total": r.get("total")}
    except (json.JSONDecodeError, OSError):
        prev_state = {}

meta = {
    "lastRefresh": str(ws["B3"].value) if ws["B3"].value else "",
    "maxPoints": ws["B4"].value,
    "participants": ws["B5"].value,
}

rows = []
row_num = 8
while True:
    player = ws.cell(row=row_num, column=2).value
    if player is None or str(player).strip() == "":
        break
    name = str(player).strip()
    rows.append({
        "rank":     ws.cell(row=row_num, column=1).value,
        "player":   name,
        "group":    ws.cell(row=row_num, column=3).value,
        "ko":       ws.cell(row=row_num, column=4).value,
        "bonus":    ws.cell(row=row_num, column=5).value,
        "total":    ws.cell(row=row_num, column=6).value,
        # "pctMax" is added later, once we know how many points are winnable so far.
    })
    row_num += 1

# Daily baseline = standings at the start of today. Movement arrows and points
# gained are measured against it, so they reflect the calendar day rather than
# each 30-min auto-sync. The baseline rolls over on the first export of a new day,
# seeded from the previous day's final standings (prev_state).
today = datetime.date.today().isoformat()
baseline = None
if os.path.exists(BASELINE_FILE):
    try:
        with open(BASELINE_FILE, encoding="utf-8") as f:
            baseline = json.load(f)
    except (json.JSONDecodeError, OSError):
        baseline = None

if not baseline or baseline.get("date") != today:
    seed = prev_state or {r["player"]: {"rank": r["rank"], "total": r["total"]} for r in rows}
    baseline = {"date": today, "players": seed}
    with open(BASELINE_FILE, "w", encoding="utf-8") as f:
        json.dump(baseline, f, ensure_ascii=False, indent=2)

base = baseline.get("players") or {}
for r in rows:
    b = base.get(r["player"])
    r["prevRank"] = b["rank"] if b else None
    r["pointsToday"] = (r["total"] - b["total"]) \
        if (b and b.get("total") is not None and r["total"] is not None) else 0

# leaderboard.json is written further down (see "finalise leaderboard.json"),
# once fixtures and the bonus answer key are known — we need them to compute how
# many points have actually been winnable so far for the % Max column.

# ───────────────── flat per-player picks table (row 100+) ─────────────────

hdr = {}
for c in range(1, ws.max_column + 1):
    h = ws.cell(row=FLAT_HEADER_ROW, column=c).value
    if h is not None:
        hdr[str(h)] = c

flat = {}                      # player name -> {header: value}
r = FLAT_HEADER_ROW + 1
while True:
    name = ws.cell(row=r, column=hdr["Name"]).value
    if name is None or str(name).strip() == "":
        break
    flat[str(name).strip()] = {h: ws.cell(row=r, column=c).value for h, c in hdr.items()}
    r += 1

# Players ordered by leaderboard rank; fall back to flat-table order.
players = [x["player"] for x in sorted(rows, key=lambda x: (x["rank"] is None, x["rank"]))
           if x["player"] in flat]
players += [n for n in flat if n not in players]

# ───────────────── Backend: fixtures, results, progression ─────────────────

# Authoritative scores + played-status come from the Results sheet input cells.
# (Backend coerces a blank result to 0, which would look like a played 0-0 draw;
# the Results F/H cells stay genuinely blank until a game is actually played.)
_rs = wb["Results"]
_GROUP_HDR = {3: "A", 12: "B", 21: "C", 30: "D", 39: "E", 48: "F",
              57: "G", 66: "H", 75: "I", 84: "J", 93: "K", 102: "L"}
results_scores = {}            # (group, frozenset({home, away})) -> (results_home, hg, ag)
for _hdr, _g in _GROUP_HDR.items():
    for _r in range(_hdr + 2, _hdr + 8):
        _h = _rs.cell(row=_r, column=3).value   # C
        _a = _rs.cell(row=_r, column=5).value   # E
        if _h and _a:
            key = (_g, frozenset({str(_h).strip(), str(_a).strip()}))
            results_scores[key] = (str(_h).strip(),
                                   as_int(_rs.cell(row=_r, column=6).value),   # F
                                   as_int(_rs.cell(row=_r, column=8).value))   # H

# Group fixtures in GS01..GS72 order (Backend rows 3-74, cols B-F) — order matters
# because it aligns positionally with the player picks GS01..GS72.
fixtures = []                  # [{group, home, away, hg, ag, played}]
for row in bk.iter_rows(min_row=3, max_row=74, min_col=2, max_col=6, values_only=True):
    grp, home, away, hg, ag = row
    if grp and home and away and len(str(grp)) == 1:
        rsc = results_scores.get((str(grp), frozenset({str(home).strip(), str(away).strip()})))
        if rsc is not None:
            rhome, rf, rh = rsc
            hg, ag = (rf, rh) if str(home).strip() == rhome else (rh, rf)
        else:
            hg, ag = as_int(hg), as_int(ag)
        fixtures.append({"group": str(grp), "home": home, "away": away,
                         "hg": hg, "ag": ag, "played": hg is not None and ag is not None})

# Kickoff datetimes from the Template sheet (same fixture order per group).
kickoffs = {}
tpl = wb["Template"] if "Template" in wb.sheetnames else None
if tpl is not None:
    import datetime as _dt
    for row in tpl.iter_rows(min_row=1, max_col=8, values_only=True):
        b, home, away = row[1], row[2], row[4]
        if isinstance(b, _dt.datetime) and home and away:
            kickoffs.setdefault((home, away), b.strftime("%Y-%m-%d %H:%M"))

# Per-group completeness flags (Backend AO/AP, rows 2-13).
group_complete = {}
for row in bk.iter_rows(min_row=2, max_row=13, min_col=41, max_col=42, values_only=True):
    g, done = row
    if g and len(str(g)) == 1:
        group_complete[str(g)] = str(done) == "True"

# Team table (Backend cols P-Z from row 3): slot, team, group + outcome flags.
teams = {}                     # team -> {"group": g, "flags": {...}}
actual_winner = {}             # group -> winning team (only if group complete)
for row in bk.iter_rows(min_row=3, max_row=50, min_col=16, max_col=26, values_only=True):
    slot, team, grp = row[0], row[1], row[2]
    if not team:
        continue
    flags = dict(zip(["R32", "R16", "QF", "SF", "Final", "Third", "RunnerUp", "Winner"],
                     [str(v) == "True" for v in row[3:11]]))
    teams[team] = {"group": str(grp) if grp else None, "flags": flags}
    if slot and str(slot).startswith("1") and len(str(slot)) == 2 and group_complete.get(str(slot)[1]):
        actual_winner[str(slot)[1]] = team

# KO rounds by match participation (Backend cols I-N from row 3).
# NOTE: the team table's "Quarter" flag column is broken in the model, so
# round membership is derived from the KO match list instead (validated).
ko_round = {"R32": set(), "R16": set(), "QF": set(), "SF": set(), "Final": set()}
round_map = {"R32": "R32", "R16": "R16", "QF": "QF", "SF": "SF", "Final": "Final"}
for row in bk.iter_rows(min_row=3, max_row=40, min_col=9, max_col=14, values_only=True):
    m, rnd, home, away, hg, ag = row
    key = round_map.get(str(rnd)) if rnd else None
    if key:
        for t in (home, away):
            if t in teams:
                ko_round[key].add(t)

champion = {t for t, d in teams.items() if d["flags"]["Winner"]}
third_team = {t for t, d in teams.items() if d["flags"]["Third"]}

def actual_stage(team):
    """Deepest stage the team has reached so far, or None before the R32 draw."""
    if team in champion:
        return "Champion"
    for key in ["Final", "SF", "QF", "R16", "R32"]:
        if team in ko_round[key]:
            return key
    if ko_round["R32"] and group_complete.get(teams[team]["group"]):
        return "Group stage"   # group done, R32 drawn, team not in it -> eliminated
    return None

# ───────────────── per-player derived picks ─────────────────

def player_round_sets(p):
    d = flat[p]
    rq  = {d.get(f"RQ_{i:02d}") for i in range(1, 33)}
    r16 = {d.get(f"R16_{i:02d}") for i in range(1, 17)}
    qf  = {d.get(f"QF_{i:02d}") for i in range(1, 9)}
    sf  = {d.get(f"SF_{i:02d}") for i in range(1, 5)}
    fin = {d.get("FIN_1"), d.get("FIN_2")}
    return {"R32": rq, "R16": r16, "QF": qf, "SF": sf, "Final": fin,
            "Winner": d.get("Winner"), "Third": d.get("Third")}

rounds_by_player = {p: player_round_sets(p) for p in players}

def predicted_stage(p, team):
    rs = rounds_by_player[p]
    if rs["Winner"] == team:
        return "Champion"
    for key in ["Final", "SF", "QF", "R16", "R32"]:
        if team in rs[key]:
            return key
    return "Group stage"

# ───────────────── group standings (from authoritative fixtures) ─────────────────
# Computed from the merged Backend/Results scores already in `fixtures`, so the
# table matches the official results. Tiebreak: Pts -> GD -> GF -> team name.
# (FIFA's head-to-head / fair-play tiebreakers are NOT applied — the feed lacks
# the data for them; flagged in the UI footnote.)

def compute_standings(group):
    tbl = {}   # team -> stat dict
    def row(t):
        return tbl.setdefault(t, {"team": t, "P": 0, "W": 0, "D": 0, "L": 0,
                                  "GF": 0, "GA": 0, "GD": 0, "Pts": 0})
    for fx in fixtures:
        if fx["group"] != group:
            continue
        h, a = row(fx["home"]), row(fx["away"])
        h["team"], a["team"]  # ensure both teams listed even if unplayed
        if not fx["played"]:
            continue
        hg, ag = fx["hg"], fx["ag"]
        h["P"] += 1; a["P"] += 1
        h["GF"] += hg; h["GA"] += ag
        a["GF"] += ag; a["GA"] += hg
        if hg > ag:
            h["W"] += 1; a["L"] += 1; h["Pts"] += 3
        elif hg < ag:
            a["W"] += 1; h["L"] += 1; a["Pts"] += 3
        else:
            h["D"] += 1; a["D"] += 1; h["Pts"] += 1; a["Pts"] += 1
    for r in tbl.values():
        r["GD"] = r["GF"] - r["GA"]
    standings = sorted(tbl.values(),
                       key=lambda r: (-r["Pts"], -r["GD"], -r["GF"], r["team"]))
    for i, r in enumerate(standings):
        r["pos"] = i + 1
    return standings


# ───────────────── games.json: scoreline distributions ─────────────────

n_players = len(players)
groups_out = {}
# Per-fixture pick popularity, keyed by fixture index, reused below to tag each
# player's own scoreline with its pool %, how many players shared it, and whether it
# was the most-picked scoreline — the substrate for the contrarian/crowd badges.
# game_scorers[gi] = how many players earned any points on that game (Giant Slayer).
game_pick_meta = {}
game_scorers = {}
for gi, fx in enumerate(fixtures):
    gs = f"GS{gi + 1:02d}"
    by_score = {}
    no_pick = []
    for p in players:
        ph, pa = as_int(flat[p].get(gs + "_H")), as_int(flat[p].get(gs + "_A"))
        if ph is None or pa is None:
            no_pick.append(p)
            continue
        by_score.setdefault((ph, pa), []).append(p)
    scorelines = []
    for (ph, pa), ppl in by_score.items():
        scorelines.append({
            "score": f"{ph}–{pa}",
            "count": len(ppl),
            "pct": round(100 * len(ppl) / n_players) if n_players else 0,
            "pts": game_points(ph, pa, fx["hg"], fx["ag"]) if fx["played"] else None,
            "players": ppl,
        })
    scorelines.sort(key=lambda s: (-s["count"], s["score"]))
    max_count = max((len(ppl) for ppl in by_score.values()), default=0)
    game_pick_meta[gi] = {
        (ph, pa): {
            "pct": round(100 * len(ppl) / n_players) if n_players else 0,
            "top": len(ppl) == max_count and max_count > 0,
            "count": len(ppl),
        }
        for (ph, pa), ppl in by_score.items()
    }
    game_scorers[gi] = (sum(len(ppl) for (ph, pa), ppl in by_score.items()
                            if game_points(ph, pa, fx["hg"], fx["ag"]) >= 1)
                        if fx["played"] else None)
    groups_out.setdefault(fx["group"], []).append({
        "kickoff": kickoffs.get((fx["home"], fx["away"]), ""),
        "home": fx["home"], "away": fx["away"],
        "played": fx["played"],
        "actual": {"hg": fx["hg"], "ag": fx["ag"]} if fx["played"] else None,
        "scorelines": scorelines,
        "noPick": len(no_pick),
    })

groups_json = []
for g in sorted(groups_out):
    winner_counts = {}
    for p in players:
        t = flat[p].get(f"GW_{g}")
        if t:
            winner_counts.setdefault(t, []).append(p)
    picks = [{
        "team": t,
        "count": len(ppl),
        "pct": round(100 * len(ppl) / n_players) if n_players else 0,
        "players": ppl,
        "correct": actual_winner.get(g) == t if g in actual_winner else None,
    } for t, ppl in winner_counts.items()]
    picks.sort(key=lambda x: (-x["count"], x["team"]))
    groups_json.append({
        "group": g,
        "complete": group_complete.get(g, False),
        "actualWinner": actual_winner.get(g),
        "winnerPicks": picks,
        "standings": compute_standings(g),
        "games": groups_out[g],
    })

with open(GAMES_OUTPUT, "w", encoding="utf-8") as f:
    json.dump({"meta": meta, "players": players, "groups": groups_json},
              f, ensure_ascii=False, indent=2)
ngames = sum(len(g["games"]) for g in groups_json)
print(f"Exported {ngames} games x {n_players} players -> {GAMES_OUTPUT}")

# ───────────────── stats.json: team progression + bonus ─────────────────

teams_json = []
for team in sorted(teams, key=lambda t: (teams[t]["group"] or "?", t)):
    dist = {s: 0 for s in STAGES}
    champions_by = []
    for p in players:
        st = predicted_stage(p, team)
        dist[st] += 1
        if st == "Champion":
            champions_by.append(p)
    teams_json.append({
        "team": team,
        "group": teams[team]["group"],
        "actual": actual_stage(team),
        "dist": dist,
        "pct": {s: (round(100 * c / n_players) if n_players else 0) for s, c in dist.items()},
        "championPct": round(100 * dist["Champion"] / n_players) if n_players else 0,
        "championBy": champions_by,
    })

# Bonus questions: texts/points from Scoring rows 19-33.
sc = wb["Scoring"]
bonus_defs = []
for row in sc.iter_rows(min_row=19, max_row=33, min_col=2, max_col=4, values_only=True):
    bid, q, pts = row
    if bid and q:
        bonus_defs.append({"id": str(bid), "q": str(q), "pts": pts})

# Answer key maintained by hand in the Results sheet: BONUS POINTS block,
# col BO = current correct answer, col BP = status. Ties are entered in BO as
# a comma-separated list ("Mexico, Canada") — every listed answer counts as
# correct. Status: "Decided" -> final (check marks shown); any other status
# with an answer present -> provisional ("so far"); empty answer -> TBD.
res = wb["Results"]
KEY_COL_ANSWER, KEY_COL_STATUS = 67, 68          # BO, BP
KEY_ROW_TO_ID = {3: "B01", 4: "B02", 5: "B03", 6: "B04", 7: "B05", 8: "B06",
                 9: "B07", 10: "B08", 11: "B09", 12: "B10", 13: "B11",
                 14: "B12", 15: "B13", 16: "B16", 17: "B15"}
STAGE_NORM = {
    "group stage": "group stage", "group": "group stage",
    "round of 32": "r32", "r32": "r32",
    "round of 16": "r16", "r16": "r16",
    "quarter-final": "qf", "quarter final": "qf", "quarter-finals": "qf", "qf": "qf",
    "semi-final": "sf", "semi final": "sf", "semi-finals": "sf", "sf": "sf",
    "final": "final",
    "winner": "champion", "champion": "champion",
}

def norm_answer(v):
    """Comparable form: stage labels unified, numbers canonical, lowercased."""
    s = str(v).strip().lower()
    if s in STAGE_NORM:
        return STAGE_NORM[s]
    try:
        return str(int(float(s.replace(",", "."))))
    except ValueError:
        return s

# Numeric questions scored with a tolerance (per the READ ME rules):
# B03 total tournament goals counts as correct within +/- 5.
NUMERIC_TOLERANCE = {"B03": 5}

# Questions credited ONLY once their answer is "decided". Most bonus questions are
# scored by direct answer-match, so the model credits a filled (even provisional)
# answer immediately (the bonus quirk) and we mirror that. B15 ("how far does
# Sweden go") is the exception: it is derived from the bracket, not a BO-match, so
# the model does NOT award it until Sweden's stage is final — crediting the
# provisional "Group stage" key early would over-count players who predicted an
# early exit. Gating it to decided keeps per-question earned reconciled with the
# model's bonus total.
DECIDED_ONLY_BONUS = {"B15"}

def hit_statuses(bid):
    """Answer-key statuses for which this question may be scored."""
    return ("decided",) if bid in DECIDED_ONLY_BONUS else ("decided", "provisional")

def answer_hit(bid, ans, key):
    """True if a player's answer matches the key (membership for ties,
    +/- tolerance for the numeric questions that allow a margin)."""
    na = norm_answer(ans)
    tol = NUMERIC_TOLERANCE.get(bid, 0)
    if tol:
        try:
            av = int(na)
            return any(abs(av - int(k)) <= tol for k in key["normSet"]
                       if k.lstrip("-").isdigit())
        except ValueError:
            pass
    return na in key["normSet"]

# Questions whose answers are bucketed into ranges for display (id -> bucket size),
# e.g. B03 total goals shown as 230-239, 240-249, … instead of every unique value.
NUMERIC_BUCKET = {"B03": 10}

def bucket_hit(lo, size, key, bid):
    """Decided-state: True if the range [lo, lo+size-1] overlaps the correct
    answer ± tolerance — i.e. a player in this bucket could be scored correct."""
    tol = NUMERIC_TOLERANCE.get(bid, 0)
    hi = lo + size - 1
    for k in key["normSet"]:
        if k.lstrip("-").isdigit() and (lo - tol) <= int(k) <= (hi + tol):
            return True
    return False

answer_key = {}
for krow, bid in KEY_ROW_TO_ID.items():
    ans = res.cell(row=krow, column=KEY_COL_ANSWER).value
    st = res.cell(row=krow, column=KEY_COL_STATUS).value
    ans_str = "" if ans is None else str(ans).strip()
    st_str = "" if st is None else str(st).strip().lower()
    if not ans_str:
        status = "tbd"
    elif st_str.startswith(("decided", "final", "klar")):
        status = "decided"
    else:
        status = "provisional"
    answer_key[bid] = {
        "current": ans_str or None,
        "status": status,
        "normSet": {norm_answer(a) for a in ans_str.replace(";", ",").split(",") if a.strip()},
    }

# ───────────── finalise leaderboard.json (now that scoring is known) ─────────────
# Bonus points don't count until the knockout stage begins. The model credits a
# filled bonus answer immediately (see the bonus scoring quirk), which would leak
# bonus points into the group-stage standings, so until the R32 draw exists we
# strip every player's bonus from their total, re-rank on the bonus-free totals,
# and exclude the bonus pool from "available". ko_round["R32"] is populated only
# once the knockout bracket has been set.
ko_started = any(ko_round[k] for k in ko_round)
if not ko_started:
    for r in rows:
        b = r.get("bonus") or 0
        if b and r.get("total") is not None:
            r["total"] = r["total"] - b
        r["bonus"] = 0
    # Re-rank (standard competition ranking — ties share a rank) and re-sort so
    # the array order matches the bonus-free standings.
    rows.sort(key=lambda x: (x["total"] is None, -(x["total"] or 0), str(x["player"])))
    last_total, last_rank = object(), 0
    for i, r in enumerate(rows, start=1):
        if r["total"] != last_total:
            last_rank, last_total = i, r["total"]
        r["rank"] = last_rank
    # Re-measure the day's movement against the (bonus-free) daily baseline.
    for r in rows:
        bse = base.get(r["player"])
        r["pointsToday"] = (r["total"] - bse["total"]) \
            if (bse and bse.get("total") is not None and r["total"] is not None) else 0

# "Available" points = the max anyone could have earned from results decided so
# far, so % Max reflects share of what was actually up for grabs (not the full
# 729-point tournament). Group games are worth 3 each once played; a bonus
# question's max counts once its answer is filled (the model credits filled
# answers immediately, decided or not) — but only from the knockout stage on.
# Knockouts are added in Phase 2.
games_played = sum(1 for fx in fixtures if fx["played"])
bonus_open = 0 if not ko_started else sum(
    (bd["pts"] or 0) for bd in bonus_defs
    if answer_key.get(bd["id"], {}).get("status", "tbd") != "tbd")
available = 3 * games_played + bonus_open
if isinstance(available, float) and available.is_integer():
    available = int(available)
meta["maxAvailable"] = available
for r in rows:
    t = r.get("total")
    r["pctMax"] = round(t / available, 4) if (available and t is not None) else 0.0

with open(OUTPUT, "w", encoding="utf-8") as f:
    json.dump({"meta": meta, "rows": rows}, f, ensure_ascii=False, indent=2)
print(f"Exported {len(rows)} players -> {OUTPUT} (available so far: {available})")

# ───────────────── history.json: daily standings (rank-over-time) ─────────────────
# One snapshot per calendar day, upserted on every export — so the latest run of
# a day overwrites that day's entry and, once the day is over, it holds that day's
# FINAL standings. Feeds the rank-trajectory sparklines and the "biggest movers
# this week" summary on the site. rank/total here are the same finalised values
# baked into leaderboard.json above (bonus-free re-rank applied pre-knockouts).
history = {"days": []}
if os.path.exists(HISTORY_OUTPUT):
    try:
        with open(HISTORY_OUTPUT, encoding="utf-8") as f:
            loaded = json.load(f)
        if isinstance(loaded, dict) and isinstance(loaded.get("days"), list):
            history = loaded
    except (json.JSONDecodeError, OSError):
        history = {"days": []}

snapshot = {
    "date": today,
    "standings": {
        r["player"]: {"rank": r["rank"], "total": r["total"]}
        for r in rows if r.get("player") is not None
    },
}
days = [d for d in history["days"] if d.get("date") != today]
days.append(snapshot)
days.sort(key=lambda d: d.get("date") or "")
history["days"] = days[-60:]          # ~2 months; the tournament runs ~6 weeks

with open(HISTORY_OUTPUT, "w", encoding="utf-8") as f:
    json.dump(history, f, ensure_ascii=False, indent=2)
print(f"Exported {len(history['days'])} day(s) of standings -> {HISTORY_OUTPUT}")


# Per-player answers per question.
def bonus_answer(p, bid):
    if bid == "B15":
        return predicted_stage(p, "Sweden")
    if bid == "B16":
        return flat[p].get("B16_TopScorer")
    return flat[p].get(bid)

bonus_json = []
for bd in bonus_defs:
    key = answer_key.get(bd["id"], {"current": None, "status": "tbd", "normSet": set()})
    counts = {}
    for p in players:
        a = bonus_answer(p, bd["id"])
        a = "—" if a is None or str(a).strip() == "" else str(a).strip()
        counts.setdefault(a, []).append(p)
    size = NUMERIC_BUCKET.get(bd["id"])
    if size:
        # Group answers into numeric ranges (e.g. 230-239) for readability.
        groups = {}                # label -> {"lo": int|None, "players": [...]}
        for a, ppl in counts.items():
            try:
                v = int(float(str(a).replace(",", ".")))
                lo = (v // size) * size
                lbl = f"{lo}–{lo + size - 1}"
            except (ValueError, TypeError):
                lo, lbl = None, str(a)        # e.g. "—" (no pick) kept as-is
            g = groups.setdefault(lbl, {"lo": lo, "players": []})
            g["players"].extend(ppl)
        tmp = []
        for lbl, g in groups.items():
            ppl = g["players"]
            tmp.append((g["lo"], {
                "answer": lbl,
                "count": len(ppl),
                "pct": round(100 * len(ppl) / n_players) if n_players else 0,
                "players": ppl,
                "hit": (bucket_hit(g["lo"], size, key, bd["id"])
                        if g["lo"] is not None
                        and key["status"] in hit_statuses(bd["id"])
                        and key["normSet"] else None),
            }))
        tmp.sort(key=lambda t: (t[0] is None, t[0] if t[0] is not None else 0))
        answers = [d for _, d in tmp]
    else:
        answers = [{
            "answer": a,
            "count": len(ppl),
            "pct": round(100 * len(ppl) / n_players) if n_players else 0,
            "players": ppl,
            "hit": answer_hit(bd["id"], a, key)
                   if (key["status"] in hit_statuses(bd["id"])
                       and key["normSet"]) else None,
        } for a, ppl in counts.items()]
        answers.sort(key=lambda x: (-x["count"], x["answer"]))
    bonus_json.append({**bd, "current": key["current"], "status": key["status"],
                       "answers": answers})

with open(STATS_OUTPUT, "w", encoding="utf-8") as f:
    json.dump({"meta": meta, "players": players, "stages": STAGES,
               "teams": teams_json, "bonus": bonus_json},
              f, ensure_ascii=False, indent=2)
print(f"Exported {len(teams_json)} teams, {len(bonus_json)} bonus questions -> {STATS_OUTPUT}")

# ───────────────── players.json: per-player detail view ─────────────────
# One record per player: their leaderboard standing plus every individual pick
# (group scorelines with points earned, group-winner picks, the knockout bracket,
# and bonus answers). Feeds the clickable player-detail panel on the site.

ko_pts_map = {"R32": 3, "R16": 6, "QF": 9, "SF": 12, "Final": 15}
lb_by_player = {r["player"]: r for r in rows}
groups_present = sorted(groups_out)

def bracket_picks(rs, key, actual_set, ptsval):
    """A player's picks for one KO round: team, whether it actually reached the
    round, and the points that pick is worth (0 until the team gets there)."""
    out = []
    for t in rs[key]:
        if not t:
            continue
        hit = t in actual_set
        out.append({"team": t, "hit": hit, "pts": ptsval if hit else 0})
    return sorted(out, key=lambda x: x["team"])

players_detail = {}
for p in players:
    d = flat[p]
    lb = lb_by_player.get(p, {})

    games = []
    for gi, fx in enumerate(fixtures):
        gs = f"GS{gi + 1:02d}"
        ph, pa = as_int(d.get(gs + "_H")), as_int(d.get(gs + "_A"))
        pts = (game_points(ph, pa, fx["hg"], fx["ag"])
               if fx["played"] and ph is not None and pa is not None else None)
        pm = (game_pick_meta.get(gi, {}).get((ph, pa))
              if ph is not None and pa is not None else None)
        games.append({
            "group": fx["group"], "home": fx["home"], "away": fx["away"],
            "ph": ph, "pa": pa,
            "played": fx["played"],
            "hg": fx["hg"] if fx["played"] else None,
            "ag": fx["ag"] if fx["played"] else None,
            "pts": pts,
            "kickoff": kickoffs.get((fx["home"], fx["away"]), ""),
            "pct": pm["pct"] if pm else None,
            "isTopPick": pm["top"] if pm else None,
            "sameCount": pm["count"] if pm else None,
            "gameScorers": game_scorers.get(gi),
        })

    gw = []
    for g in groups_present:
        t = d.get(f"GW_{g}")
        if t:
            gw.append({"group": g, "team": t,
                       "correct": (actual_winner.get(g) == t) if g in actual_winner else None})

    rs = rounds_by_player[p]
    bracket = {
        "R32":   bracket_picks(rs, "R32", ko_round["R32"], 3),
        "R16":   bracket_picks(rs, "R16", ko_round["R16"], 6),
        "QF":    bracket_picks(rs, "QF", ko_round["QF"], 9),
        "SF":    bracket_picks(rs, "SF", ko_round["SF"], 12),
        "Final": bracket_picks(rs, "Final", ko_round["Final"], 15),
    }
    win, third = rs["Winner"], rs["Third"]
    bracket["winner"] = ({"team": win, "hit": win in champion,
                          "pts": 30 if win in champion else 0} if win else None)
    bracket["third"] = ({"team": third, "hit": third in third_team,
                         "pts": 15 if third in third_team else 0} if third else None)

    bonus_ans = []
    for bd in bonus_defs:
        key = answer_key.get(bd["id"], {"status": "tbd", "normSet": set()})
        a = bonus_answer(p, bd["id"])
        a = None if a is None or str(a).strip() == "" else str(a).strip()
        # Score both decided AND provisional answers (the model credits points the
        # moment an answer is filled, decided or not), so the UI can show a green
        # check once final and a yellow check while points are still provisional.
        hit = (answer_hit(bd["id"], a, key)
               if (a and key["status"] in hit_statuses(bd["id"])
                   and key["normSet"]) else None)
        # Points this question is currently contributing to the player's bonus
        # total. The model credits a filled answer on match (decided or not), so
        # earned == pts whenever the answer hits; summed over a player's
        # questions it reconciles with their Bonus total.
        earned = bd["pts"] if hit else 0
        bonus_ans.append({"id": bd["id"], "q": bd["q"], "answer": a,
                          "hit": hit, "pts": bd["pts"], "status": key["status"],
                          "earned": earned})

    players_detail[p] = {
        "rank": lb.get("rank"), "total": lb.get("total"),
        "group": lb.get("group"), "ko": lb.get("ko"), "bonus": lb.get("bonus"),
        "pctMax": lb.get("pctMax"), "pointsToday": lb.get("pointsToday"),
        "games": games, "gw": gw, "bracket": bracket, "bonusAns": bonus_ans,
    }

with open(PLAYERS_OUTPUT, "w", encoding="utf-8") as f:
    json.dump({"meta": meta, "players": players, "detail": players_detail},
              f, ensure_ascii=False, indent=2)
print(f"Exported detail for {len(players_detail)} players -> {PLAYERS_OUTPUT}")

# ───────────────── achievements.json: gamification badges ─────────────────
# Tongue-in-cheek group-stage badges, computed entirely from data already built
# above: per-player games (with pts + pick popularity), the leaderboard rows, and
# the daily rank history. Output is a static catalog (all 20, for the locked/earned
# gallery), an earners-count per badge, and the unlocked ids per player.
ACHIEVEMENTS = [
    {"id": "crystal-ball", "name": "Crystal Ball", "emoji": "🔮", "rarity": "Legendary",
     "desc": "The only player to nail a match's exact scoreline."},
    {"id": "untouchable", "name": "Untouchable", "emoji": "👑", "rarity": "Legendary",
     "desc": "Held rank #1 three update days in a row."},
    {"id": "the-sweep", "name": "The Sweep", "emoji": "🧹", "rarity": "Legendary",
     "desc": "Got the result right on every match of a full group."},
    {"id": "perfect-day", "name": "Perfect Day", "emoji": "💯", "rarity": "Legendary",
     "desc": "Scored on every match you picked on a matchday."},
    {"id": "giant-slayer", "name": "Giant Slayer", "emoji": "🗡️", "rarity": "Rare",
     "desc": "Scored on a game only 3 or fewer players read right."},
    {"id": "hot-streak", "name": "Hot Streak", "emoji": "🔥", "rarity": "Rare",
     "desc": "Scored points in five matches running."},
    {"id": "the-comeback", "name": "The Comeback", "emoji": "📈", "rarity": "Rare",
     "desc": "Climbed 8+ ranks in a single day."},
    {"id": "day-winner", "name": "Day Winner", "emoji": "⏱️", "rarity": "Rare",
     "desc": "Top points-scorer of a matchday."},
    {"id": "contrarian-king", "name": "Contrarian King", "emoji": "🐴", "rarity": "Rare",
     "desc": "Won the day while picking against the crowd."},
    {"id": "on-the-rise", "name": "On the Rise", "emoji": "🧗", "rarity": "Rare",
     "desc": "Improved your rank three update days in a row."},
    {"id": "bullseye", "name": "Bullseye", "emoji": "🎯", "rarity": "Rare",
     "desc": "Two exact scorelines in one matchday."},
    {"id": "off-the-mark", "name": "Off the Mark", "emoji": "✅", "rarity": "Common",
     "desc": "Got your first points on the board."},
    {"id": "first-blood", "name": "First Blood", "emoji": "🩸", "rarity": "Common",
     "desc": "Landed your first exact scoreline."},
    {"id": "crowd-surfer", "name": "Crowd Surfer", "emoji": "🐑", "rarity": "Common",
     "desc": "Cashed in on the people's pick."},
    {"id": "reading-the-game", "name": "Reading the Game", "emoji": "🧠", "rarity": "Common",
     "desc": "Three correct results in a single matchday."},
    {"id": "steady-hand", "name": "Steady Hand", "emoji": "📏", "rarity": "Common",
     "desc": "Scored points on three matchdays in a row."},
    {"id": "wooden-spoon", "name": "Wooden Spoon", "emoji": "🥄", "rarity": "Common",
     "desc": "Propping up the table."},
    {"id": "snake-eyes", "name": "Snake Eyes", "emoji": "🎲", "rarity": "Common",
     "desc": "Five matches in a row without a single point."},
    {"id": "cold-snap", "name": "Cold Snap", "emoji": "🧊", "rarity": "Common",
     "desc": "Blanked an entire matchday."},
    {"id": "free-fall", "name": "Free Fall", "emoji": "📉", "rarity": "Common",
     "desc": "Tumbled 8+ ranks in a single day."},
]
ACH_ORDER = {a["id"]: i for i, a in enumerate(ACHIEVEMENTS)}


def _run_end(flags, k):
    """Index at which the first run of >=k consecutive truthy flags completes, else None."""
    run = 0
    for i, f in enumerate(flags):
        run = run + 1 if f else 0
        if run >= k:
            return i
    return None


def _gdate(g):
    return (g.get("kickoff") or "")[:10]


def _gmatch(g):
    return f'{g["home"]} v {g["away"]}'


# Worst (highest) rank on the board → Wooden Spoon.
_ranks = [r["rank"] for r in rows if r.get("rank") is not None]
worst_rank = max(_ranks) if _ranks else None
_refresh_date = str(meta.get("lastRefresh") or "")[:10] or None

# Per-player rank trajectory across update days as (date, rank) pairs. Skip days where
# the whole field is tied at rank 1 (the opening 0-0 day before any game is scored) —
# drops/climbs measured off that artificial start aren't real movement.
rank_series = {}
for _day in sorted(history.get("days", []), key=lambda x: x.get("date") or ""):
    _standings = _day.get("standings") or {}
    _dranks = [s.get("rank") for s in _standings.values() if s.get("rank") is not None]
    if _dranks and max(_dranks) == 1:
        continue
    for _pl, _st in _standings.items():
        if _st.get("rank") is not None:
            rank_series.setdefault(_pl, []).append((_day.get("date"), _st["rank"]))

# Cross-player matchday points: date -> player -> {pts, scoring_pcts}.
day_points = {}
for _p in players:
    for _g in players_detail[_p]["games"]:
        if not _g["played"] or _g["pts"] is None:
            continue
        _date = (_g.get("kickoff") or "")[:10]
        if not _date:
            continue
        _rec = day_points.setdefault(_date, {}).setdefault(_p, {"pts": 0, "scoring_pcts": []})
        _rec["pts"] += _g["pts"]
        if _g["pts"] >= 1 and _g.get("pct") is not None:
            _rec["scoring_pcts"].append(_g["pct"])

# Sole top scorer of each matchday (must have outscored everyone else, > 0). Ties
# award no Day Winner — early one-game days produce big ties and would otherwise hand
# the badge to almost the whole pool, defeating its rarity.
day_top = {}
for _date, _plm in day_points.items():
    _mx = max((v["pts"] for v in _plm.values()), default=0)
    _leaders = [pl for pl, v in _plm.items() if v["pts"] == _mx]
    if _mx > 0 and len(_leaders) == 1:
        day_top[_date] = set(_leaders)


def player_badges(p):
    """Return {badge_id: {"date": iso, "how": text}} — the first time each badge's
    condition was met, with the triggering evidence for the player profile."""
    d = players_detail[p]
    awards = {}

    def give(bid, date, how):
        if bid not in awards:        # first occurrence wins
            awards[bid] = {"date": date, "how": how}

    played = [g for g in d["games"] if g["played"] and g["pts"] is not None]
    chrono = sorted(played, key=lambda g: (g.get("kickoff") or "", g["group"], g["home"]))

    by_day = {}
    for g in played:
        dt = _gdate(g)
        if dt:
            by_day.setdefault(dt, []).append(g)

    # ── single-game (chronological → first occurrence) ──
    for g in chrono:
        if g["pts"] >= 1:
            give("off-the-mark", _gdate(g), f'First points: {_gmatch(g)} (+{g["pts"]})')
            break
    for g in chrono:
        if g["pts"] == 3:
            give("first-blood", _gdate(g), f'{g["home"]} {g["ph"]}–{g["pa"]} {g["away"]} — exact!')
            if g.get("sameCount") == 1:
                give("crystal-ball", _gdate(g),
                     f'{g["home"]} {g["ph"]}–{g["pa"]} {g["away"]} — the only exact call')
        if g["pts"] >= 1:
            gs = g.get("gameScorers")
            if gs is not None and gs <= 3:
                give("giant-slayer", _gdate(g),
                     f'{_gmatch(g)} — only {gs} player{"" if gs == 1 else "s"} scored')
            if g.get("isTopPick"):
                give("crowd-surfer", _gdate(g),
                     f'{_gmatch(g)} — rode the {g["ph"]}–{g["pa"]} consensus')

    # ── matchday-based (chronological dates → first) ──
    for dt in sorted(by_day):
        dg = by_day[dt]
        if len(dg) >= 2 and all(x["pts"] >= 1 for x in dg):
            give("perfect-day", dt, f'Scored on all {len(dg)} games that day')
        n3 = sum(1 for x in dg if x["pts"] == 3)
        if n3 >= 2:
            give("bullseye", dt, f'{n3} exact scorelines in a day')
        nr = sum(1 for x in dg if x["pts"] >= 1)
        if nr >= 3:
            give("reading-the-game", dt, f'{nr} correct results in a day')
        if len(dg) >= 2 and all(x["pts"] == 0 for x in dg):
            give("cold-snap", dt, f'Blanked all {len(dg)} games that day')

    # ── group sweep: every match of a fully-played group correct ──
    by_group = {}
    for g in played:
        by_group.setdefault(g["group"], []).append(g)
    for grp, gg in sorted(by_group.items()):
        if len(gg) >= 6 and all(x["pts"] >= 1 for x in gg):
            give("the-sweep", max(_gdate(x) for x in gg), f'All 6 Group {grp} results correct')

    # ── streaks (chronological) ──
    i = _run_end([g["pts"] >= 1 for g in chrono], 5)
    if i is not None:
        give("hot-streak", _gdate(chrono[i]), "Points in 5 straight matches")
    i = _run_end([g["pts"] == 0 for g in chrono], 5)
    if i is not None:
        give("snake-eyes", _gdate(chrono[i]), "0 points in 5 straight matches")

    # ── day winner / contrarian (cross-player) ──
    for dt in sorted(day_top):
        if p in day_top[dt]:
            give("day-winner", dt, f'Outright top scorer (+{day_points[dt][p]["pts"]})')
            pcts = day_points[dt][p]["scoring_pcts"]
            if pcts and (sum(pcts) / len(pcts)) < 25:
                give("contrarian-king", dt,
                     f'Top scorer on a {round(sum(pcts) / len(pcts))}%-popularity card')

    # ── rank movement / consistency (history) ──
    series = rank_series.get(p, [])
    for (d0, r0), (d1, r1) in zip(series, series[1:]):
        if r0 - r1 >= 8:
            give("the-comeback", d1, f'#{r0} → #{r1} in a day')
        if r1 - r0 >= 8:
            give("free-fall", d1, f'#{r0} → #{r1} in a day')
    for j in range(len(series) - 2):
        (da, ra), (db, rb), (dc, rc) = series[j], series[j + 1], series[j + 2]
        if ra == 1 and rb == 1 and rc == 1:
            give("untouchable", dc, "Held #1 three update days running")
        if ra > rb > rc:
            give("on-the-rise", dc, f'#{ra} → #{rb} → #{rc}')

    # ── points on three consecutive matchdays ──
    dts = sorted(by_day)
    flags = [sum(x["pts"] for x in by_day[dt]) > 0 for dt in dts]
    e = _run_end(flags, 3)
    if e is not None:
        give("steady-hand", dts[e], "Points on 3 matchdays running")

    # ── wooden spoon (current standing) ──
    if worst_rank is not None and d.get("rank") == worst_rank:
        give("wooden-spoon", _refresh_date, f'Last place — rank {d.get("rank")}')

    return awards


by_player = {}
rarity_count = {a["id"]: 0 for a in ACHIEVEMENTS}
for p in players:
    awards = player_badges(p)
    by_player[p] = [{"id": bid, "date": awards[bid]["date"], "how": awards[bid]["how"]}
                    for bid in sorted(awards, key=lambda bid: ACH_ORDER[bid])]
    for bid in awards:
        rarity_count[bid] += 1

with open(ACH_OUTPUT, "w", encoding="utf-8") as f:
    json.dump({"meta": meta, "catalog": ACHIEVEMENTS,
               "rarityCount": rarity_count, "byPlayer": by_player},
              f, ensure_ascii=False, indent=2)
_total_badges = sum(len(v) for v in by_player.values())
print(f"Exported {_total_badges} badges across {len(by_player)} players -> {ACH_OUTPUT}")

# ───────────────── bracket.json: knockout bracket ─────────────────
# Reads the Backend KO match list (M73..M104, cols I-N) — the fixed FIFA 2026
# bracket. M73-M88 = R32, M89-M96 = R16, M97-M100 = QF, M101-M102 = SF,
# M103 = third-place (Bronze), M104 = Final. The tree is wired by sequential
# pairing (winners of M73,M74 -> M89, etc.), which is the standard bracket order.
# Home/Away are "TBD"/0 until the R32 draw is made, so before the draw this
# feed is the empty structure (every slot a placeholder) — rendered as such.

# Map each match to its round and its left/right half (Final/Bronze are centre).
KO_LAYOUT = {}   # num -> (round, side)
for _n in range(73, 89):  KO_LAYOUT[_n] = ("R32", "L" if _n <= 80 else "R")
for _n in range(89, 97):  KO_LAYOUT[_n] = ("R16", "L" if _n <= 92 else "R")
for _n in range(97, 101): KO_LAYOUT[_n] = ("QF",  "L" if _n <= 98 else "R")
KO_LAYOUT[101] = ("SF", "L"); KO_LAYOUT[102] = ("SF", "R")
KO_LAYOUT[103] = ("Bronze", "C"); KO_LAYOUT[104] = ("Final", "C")

# Feeder children: which two earlier matches' winners meet in this match.
KO_FEEDERS = {}
for _k in range(8):  KO_FEEDERS[89 + _k] = (73 + 2 * _k, 74 + 2 * _k)   # R16 <- R32 pairs
for _k in range(4):  KO_FEEDERS[97 + _k] = (89 + 2 * _k, 90 + 2 * _k)   # QF  <- R16 pairs
KO_FEEDERS[101] = (97, 98); KO_FEEDERS[102] = (99, 100)                 # SF  <- QF pairs
KO_FEEDERS[104] = (101, 102)                                           # Final <- SF winners
KO_FEEDERS[103] = (101, 102)                                           # Bronze <- SF losers

NEXT_ROUND = {"R32": "R16", "R16": "QF", "QF": "SF", "SF": "Final", "Final": "Champion"}

def real_team(v):
    """Backend uses 'TBD' or 0 for an unfilled slot; only a known team is real."""
    return v if (v and v in teams) else None

def advance_pct(team, target):
    """% of the pool that predicted `team` to reach `target` (the round this
    match feeds into). For the Final, `target` is winning it (Champion)."""
    if not team or not n_players:
        return None
    if target == "Champion":
        c = sum(1 for p in players if rounds_by_player[p]["Winner"] == team)
    else:
        c = sum(1 for p in players if team in rounds_by_player[p].get(target, set()))
    return {"pct": round(100 * c / n_players), "count": c}

ko_matches = []
for row in bk.iter_rows(min_row=3, max_row=40, min_col=9, max_col=14, values_only=True):
    mid, rnd, home, away, hg, ag = row
    if not mid:
        continue
    num = as_int(str(mid).lstrip("Mm"))
    if num not in KO_LAYOUT:
        continue
    rnd_key, side = KO_LAYOUT[num]
    h, a = real_team(home), real_team(away)
    hg, ag = as_int(hg), as_int(ag)
    # Winner is taken from the model's authoritative progression sets, not the
    # raw score (a KO tie is decided on penalties, which the score may not show).
    winner = None
    if h and a:
        if rnd_key == "Final":
            winner = h if h in champion else (a if a in champion else None)
        else:
            nxt = ko_round.get(NEXT_ROUND.get(rnd_key), set())
            if rnd_key == "SF":   # SF winners are the Final's two participants
                nxt = ko_round.get("Final", set())
            winner = h if h in nxt else (a if a in nxt else None)
    has_score = (hg is not None and ag is not None and (hg or ag))
    target = NEXT_ROUND.get(rnd_key)
    ko_matches.append({
        "id": f"M{num}", "no": num, "round": rnd_key, "side": side,
        "feeders": [f"M{c}" for c in KO_FEEDERS.get(num, ())],
        "home": h, "away": a,
        "hg": hg if (h and a and has_score) else None,
        "ag": ag if (h and a and has_score) else None,
        "winner": winner,
        "homePick": advance_pct(h, target) if h else None,
        "awayPick": advance_pct(a, target) if a else None,
    })

drawn = any(m["home"] or m["away"] for m in ko_matches)
with open(BRACKET_OUTPUT, "w", encoding="utf-8") as f:
    json.dump({"meta": {**meta, "drawn": drawn}, "matches": ko_matches},
              f, ensure_ascii=False, indent=2)
print(f"Exported {len(ko_matches)} knockout matches (drawn={drawn}) -> {BRACKET_OUTPUT}")

# ───────────────── self-check vs official leaderboard ─────────────────

# The model gates ALL group-winner points (Scoring!S) on Backend!AL11
# ("group stage complete?") — they stay 0 until every group has finished,
# then all 12 are credited together. Mirror that here so the self-check
# doesn't flag the transient window where some groups are done but others
# aren't (otherwise it warns by 3 pts per correct winner of a finished group).
group_stage_complete = all(group_complete.get(g) for g in "ABCDEFGHIJKL")

for lbrow in rows:
    p = lbrow["player"]
    if p not in flat:
        continue
    gpts = 0
    for gi, fx in enumerate(fixtures):
        if not fx["played"]:
            continue
        ph = as_int(flat[p].get(f"GS{gi + 1:02d}_H"))
        pa = as_int(flat[p].get(f"GS{gi + 1:02d}_A"))
        if ph is not None and pa is not None:
            gpts += game_points(ph, pa, fx["hg"], fx["ag"])
    if group_stage_complete:
        for g, w in actual_winner.items():
            if flat[p].get(f"GW_{g}") == w:
                gpts += 3
    rs = rounds_by_player[p]
    kpts = sum(len(rs[k] & ko_round[k]) * v for k, v in ko_pts_map.items())
    kpts += 30 if rs["Winner"] in champion else 0
    kpts += 15 if rs["Third"] in third_team else 0
    if lbrow["group"] is not None and gpts != lbrow["group"]:
        print(f"WARNING: {p} computed group pts {gpts} != leaderboard {lbrow['group']}")
    if lbrow["ko"] is not None and kpts != lbrow["ko"]:
        print(f"WARNING: {p} computed KO pts {kpts} != leaderboard {lbrow['ko']}")

# ── OG social-preview image (og-image.png) ───────────────────────────────────
OG_OUTPUT = "og-image.png"
try:
    from PIL import Image, ImageDraw, ImageFont as _IF
    # Brand the share image from config.json so each pool gets its own name,
    # colours and URL (falls back to OPX navy if config is missing). Keeps a
    # single export.py syncable across all 4 pools — only WORKBOOK differs.
    def _hex_rgb(_h, _default):
        try:
            _h = _h.lstrip("#")
            return (int(_h[0:2], 16), int(_h[2:4], 16), int(_h[4:6], 16))
        except Exception:
            return _default
    _cfg = {}
    try:
        with open("config.json", encoding="utf-8") as _cf:
            _cfg = json.load(_cf)
    except Exception:
        pass
    _pool  = _cfg.get("poolName", "OPX")
    _c_top = _hex_rgb(_cfg.get("brandColor", ""),     (26, 37, 96))
    _c_bot = _hex_rgb(_cfg.get("brandColorDark", ""), (20, 42, 108))
    _url   = (_cfg.get("siteUrl", "https://andreaspanagos.github.io/WC-2026-OPX-leaderboard/")
              .replace("https://", "").replace("http://", "").rstrip("/"))
    _W, _H = 1200, 630
    img = Image.new("RGB", (_W, _H))
    draw = ImageDraw.Draw(img)
    # Brand-colour vertical gradient (brandColor -> brandColorDark)
    for _y in range(_H):
        _t = _y / _H
        draw.rectangle([0, _y, _W, _y + 1], fill=(
            int(_c_top[0] + _t * (_c_bot[0] - _c_top[0])),
            int(_c_top[1] + _t * (_c_bot[1] - _c_top[1])),
            int(_c_top[2] + _t * (_c_bot[2] - _c_top[2])),
        ))
    # Gold accent bar at top
    draw.rectangle([0, 0, _W, 10], fill=(255, 215, 0))
    try:
        _fp = "C:/Windows/Fonts/"
        _big   = _IF.truetype(_fp + "arialbd.ttf", 72)
        _med   = _IF.truetype(_fp + "arialbd.ttf", 44)
        _reg   = _IF.truetype(_fp + "arial.ttf",   30)
        _small = _IF.truetype(_fp + "arial.ttf",   22)
    except OSError:
        _big = _med = _reg = _small = _IF.load_default()
    draw.text((80,  38), "FIFA World Cup 2026",   font=_reg,  fill=(180, 200, 255))
    draw.text((80,  78), f"{_pool} Live Leaderboard",  font=_big,  fill=(255, 215, 0))
    draw.rectangle([80, 178, _W - 80, 181],        fill=(55, 72, 130))
    draw.text((80, 192), f"Updated: {today}",      font=_small, fill=(140, 165, 220))
    for _i, _r in enumerate(rows[:3]):
        _col = [(255, 215, 0), (200, 200, 200), (200, 135, 65)][_i]
        _y2  = 258 + _i * 108
        draw.text((80,  _y2), f"#{_i + 1}", font=_med, fill=_col)
        draw.text((168, _y2), str(_r.get("player", ""))[:22], font=_med, fill=(255, 255, 255))
        draw.text((168, _y2 + 50), f"{_r.get('total', 0)} pts", font=_reg, fill=(155, 185, 235))
    draw.rectangle([0, 600, _W, 630], fill=(12, 18, 55))
    draw.text((80, 607), _url,
              font=_small, fill=(110, 145, 200))
    img.save(OG_OUTPUT, "PNG")
    print(f"Generated {OG_OUTPUT}")
except ImportError:
    print(f"Pillow not found — skipping {OG_OUTPUT}  (pip install Pillow to enable)")
except Exception as _og_err:
    print(f"OG image skipped: {_og_err}")
